import calendar
import format_report
from dataclasses import dataclass
from  openpyxl import load_workbook

@dataclass
class Delay:
    Operational_Date    : str
    Time                : str
    Event               : str
    Category_Code       : str
    Schedule_Arrival    : str
    Actual_Arrival      : str
    Delay_Code          : str
    Primary_Incident    : str
    Train               : str
    Consist             : str
    Car                 : str
    Route               : str
    Location            : str
    Dep                 : str
    Arr                 : str
    Status              : str
    blank               : str
    Remarks             : str
    Remarks_Added_By    : str
    Event_Creation_Time : str

def break_down_date(date_string):
    year = date_string[:4]
    month = int(date_string[5:7])
    full_month_name = calendar.month_name[month]
    day = date_string[-2:]

    return year, full_month_name, day

def report_to_delay_list(file, filter_codes):
    delay_list = []
    wb = load_workbook(file)
    ws = wb.active
    max_row = ws.max_row
    delay_code_column = 'G'

    for active_row in range(4, max_row + 1):
        if ws[f'{delay_code_column}{active_row}'].value in filter_codes:

            delay_list.append(Delay(ws.cell(row = active_row, column = 1 ).value,
                                    ws.cell(row = active_row, column = 2 ).value,
                                    ws.cell(row = active_row, column = 3 ).value,
                                    ws.cell(row = active_row, column = 4 ).value,
                                    ws.cell(row = active_row, column = 5 ).value,
                                    ws.cell(row = active_row, column = 6 ).value,
                                    ws.cell(row = active_row, column = 7 ).value,
                                    ws.cell(row = active_row, column = 8 ).value,
                                    ws.cell(row = active_row, column = 9 ).value,
                                    ws.cell(row = active_row, column = 10).value,
                                    ws.cell(row = active_row, column = 11).value,
                                    ws.cell(row = active_row, column = 12).value,
                                    ws.cell(row = active_row, column = 13).value,
                                    ws.cell(row = active_row, column = 14).value,
                                    ws.cell(row = active_row, column = 15).value,
                                    ws.cell(row = active_row, column = 16).value,
                                    ws.cell(row = active_row, column = 17).value,
                                    ws.cell(row = active_row, column = 18).value,
                                    ws.cell(row = active_row, column = 19).value,
                                    ws.cell(row = active_row, column = 20).value,))

    wb.close
    return delay_list

def get_dates (delay_list):
    all_dates= []
    for delay in delay_list:
        all_dates.append(delay.Operational_Date)
    return list(set(all_dates))

def get_delays(list, date, codes):
    filtered_delays = []
    for delay in list:
        if delay.Delay_Code in codes and delay.Operational_Date == date:
            filtered_delays.append(delay)
    return filtered_delays

def sort_delays(list):
    operational_day = '03:00'
    sorted_delays = sorted(list, key=lambda Delay: Delay.Time)
    carryover_delays = []
    delays = []
    for delay in sorted_delays:
        if delay.Time == '00:00': delay.Time = '00:01'
        
        if delay.Time >= '00:01' and delay.Time < operational_day:
            carryover_delays.append(delay)
        if delay.Time >= operational_day and delay.Time <= '23:59':
            delays.append(delay)

    if len(carryover_delays) == 0:
        return delays
    else:
        return delays + carryover_delays
    
def process_report(downloaded_file):

    BT_delay_codes = ['BTCC', 'BTCT', 'BTDV', 'BTFP', 'BTKP', 'BTKQ', 'BTMN', 'BTOP', 'BTRV', 'BTSF',
                    'BTUP', 'BTWA', 'BTWC', 'BTWM', 'BTWR'] 

    QQ_delay_codes = ['QQAC', 'QQAF', 'QQDF', 'QQEF', 'QQEX', 'QQFI', 'QQHB', 'QQLA', 'QQME',
                    'QQMQ', 'QQOL', 'QQPF', 'QQRF', 'QQSF', 'QQTE', 'QQWC', 'QQWN', 'QQWR',
                    'QQWS']
    
    Alstom_delay_codes = BT_delay_codes + QQ_delay_codes

    all_Alstom_delays_list = report_to_delay_list(downloaded_file, Alstom_delay_codes)
    
    dates_in_report = get_dates(all_Alstom_delays_list)
    dates_in_report.sort()

    generated_file_names = []
    dates_without_delays = []


    for date in dates_in_report:
        year, full_month_name, day = break_down_date(date)
        
        bt_delays = get_delays(all_Alstom_delays_list, date, BT_delay_codes )
        qq_delays = get_delays(all_Alstom_delays_list, date, QQ_delay_codes )

        if len(bt_delays) > 0:
            bt_delays = sort_delays(bt_delays)
            filename = f'Atlas - L102 - Delay to Train Details {full_month_name} {day} BT.xlsx'
       
            format_report.format(bt_delays, filename)
            generated_file_names.append(filename)
        else:
            dates_without_delays.append(f'No BT delays on {full_month_name} {day}')

        if len(qq_delays) > 0:
            qq_delays = sort_delays(qq_delays)
            filename = f'Atlas - L102 - Delay to Train Details {full_month_name} {day} QQ.xlsx'
       
            format_report.format(qq_delays, filename)
            generated_file_names.append(filename)
        else:
            dates_without_delays.append(f'No QQ delays on {full_month_name} {day}')
        
    return generated_file_names, dates_without_delays
 
    

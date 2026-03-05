from openpyxl import load_workbook
from openpyxl.styles import Alignment


def format(delay_list, filename):
    template_file = 'Delay to Train Report REV1.xlsx'

    wb = load_workbook(template_file)
    ws = wb.active

    
    row_number  = 3

    
    for delay in delay_list:
        ws[f'B{row_number}'] = delay.Operational_Date
        ws[f'C{row_number}'] = delay.Time
        ws[f'D{row_number}'] = delay.Event
        ws[f'E{row_number}'] = delay.Category_Code
        ws[f'F{row_number}'] = delay.Schedule_Arrival
        ws[f'G{row_number}'] = delay.Actual_Arrival
        ws[f'H{row_number}'] = delay.Delay_Code
        ws[f'I{row_number}'] = delay.Primary_Incident
        ws[f'J{row_number}'] = delay.Train
        ws[f'K{row_number}'] = delay.Consist
        ws[f'L{row_number}'] = delay.Car
        ws[f'M{row_number}'] = delay.Route
        ws[f'N{row_number}'] = delay.Location
        ws[f'O{row_number}'] = delay.Dep
        ws[f'P{row_number}'] = delay.Arr
        ws[f'Q{row_number}'] = delay.Status
        ws[f'S{row_number}'] = delay.Remarks

        
        ws[f'S{row_number}'].alignment = Alignment(wrap_text=True)
        ws[f'T{row_number}'] = delay.Remarks_Added_By
        ws[f'U{row_number}'] = delay.Event_Creation_Time
        
        ws.row_dimensions[row_number].height = None
        
        row_number += 1
        
        
    ws.column_dimensions['U'].width = 21
    ws.column_dimensions['Q'].width = 12.11
    wb.save(filename)